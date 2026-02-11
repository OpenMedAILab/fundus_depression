import re
import os
import json
import pandas as pd
from tqdm import tqdm
import shutil
import numpy as np
from openpyxl import load_workbook
import random
import imghdr


data_root_dir = 'fundus_depression/ukb_dataset'
img_root_dir = os.path.join(data_root_dir, 'images')
data_split_result_dir = os.path.join(data_root_dir, 'data_splits')


label_excel_file = os.path.join(data_root_dir, 'v2-0820_ukb_depression_fundus.xlsx')
image_path_prefix = 'depression_dataset'
TOTAL_N = 43445  # 指定的总样本数
SEED = 42

# 在文件顶部或合适位置添加这个函数
def is_image_file(fpath: str) -> bool:
    """
    判断文件是否为图片：优先使用 python-magic（libmagic），然后回退到 imghdr，再回退到 Pillow verify。
    如果文件不存在或不可读返回 False。
    """
    if not os.path.isfile(fpath):
        return False

    # 尝试 python-magic
    try:
        import magic  # type: ignore
        try:
            mime = magic.from_file(fpath, mime=True)
            if isinstance(mime, bytes):
                mime = mime.decode('utf-8', errors='ignore')
            if mime and str(mime).startswith('image/'):
                return True
        except Exception:
            pass
    except Exception:
        pass

    # 回退到 imghdr
    try:
        if imghdr.what(fpath) is not None:
            return True
    except Exception:
        pass

    # 最后回退到 Pillow verify（如果安装）
    try:
        from PIL import Image  # type: ignore
        try:
            with Image.open(fpath) as im:
                im.verify()
            return True
        except Exception:
            return False
    except Exception:
        return False


def move_non_images_to_trash(img_dir: str, show_progress: bool = True):
    trash_root = img_dir + '_trash'
    os.makedirs(trash_root, exist_ok=True)


    # 收集待检查文件（仅 left/right）
    files_to_check = []
    for sub in ('left', 'right'):
        sub_dir = os.path.join(img_dir, sub)
        if not os.path.isdir(sub_dir):
            continue
        for root, dirs, files in os.walk(sub_dir):
            # 跳过已在 trash 下的路径
            if os.path.abspath(root).startswith(os.path.abspath(trash_root)):
                continue
            for fname in files:
                files_to_check.append((root, fname))

    moved = 0
    checked = 0

    iterator = files_to_check
    if show_progress:
        iterator = tqdm(files_to_check, desc="Checking images", unit="file")

    for root, fname in iterator:
        checked += 1
        fpath = os.path.join(root, fname)
        is_image = is_image_file(fpath)

        if not is_image:
            rel_dir = os.path.relpath(root, img_dir)
            dest_dir = os.path.join(trash_root, rel_dir) if rel_dir != '.' else trash_root
            os.makedirs(dest_dir, exist_ok=True)
            dest_path = os.path.join(dest_dir, fname)
            try:
                shutil.move(fpath, dest_path)
                moved += 1
            except Exception as e:
                print(f"移动文件失败: {fpath} -> {dest_path}，原因: {e}")

        if show_progress and hasattr(iterator, "set_postfix"):
            iterator.set_postfix({'checked': checked, 'moved': moved})

    if show_progress and hasattr(iterator, "close"):
        try:
            iterator.close()
        except Exception:
            pass

    print(f"检查完成，总文件数: {checked}，已移动非图片文件: {moved} 到 `{trash_root}`")
    return moved

# move_non_images_to_trash(img_root_dir)


def make_serializable(obj):
    if isinstance(obj, dict):
        return {k: make_serializable(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [make_serializable(v) for v in obj]
    # pandas / numpy 缺失值
    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass
    # bytes -> str（先解码以便后续字符串映射）
    if isinstance(obj, (bytes, bytearray)):
        try:
            obj = obj.decode('utf-8')
        except Exception:
            obj = str(obj)
    # 字符串映射：'否' -> 0, '是' -> 1
    if isinstance(obj, str):
        s = obj.strip()
        if s == '否':
            return 0
        if s == '是':
            return 1
        return s
    # pandas Timestamp -> ISO 字符串
    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()
    # pandas Timedelta -> 字符串
    if isinstance(obj, pd.Timedelta):
        return str(obj)
    # numpy 标量 -> Python 原生
    if isinstance(obj, np.generic):
        return obj.item()
    return obj

def build_ukb_dataset(label_excel_file: str,
                      img_dir: str,
                      save: bool = True):
    """
    构建 ukb_dataset 并返回 (ukb_dataset, missing_images)。
    如果 save 为 True，会写入 `ukb_dataset.json` 和 `missing_images.txt` 到 data_root_dir。
    假设模块中已定义 `make_serializable(obj)`。
    """
    df = pd.read_excel(label_excel_file)
    df['eid_ckd'] = df['eid_ckd'].astype(str)

    ukb_dataset = {}
    missing_images = []

    for subfolder in tqdm(os.listdir(img_dir), desc="Processing subfolders"):
        subfolder_path = os.path.join(img_dir, subfolder)
        if not os.path.isdir(subfolder_path) or subfolder not in ['left', 'right']:
            continue

        image_names = os.listdir(subfolder_path)
        txt_file_path = os.path.join(img_dir, f"{subfolder}_images.txt")
        with open(txt_file_path, 'w') as txt_file:
            txt_file.write('\n'.join(image_names))

        no_clinical_dir = subfolder_path + '_no_clinical_information'


        for image_name in tqdm(image_names, desc=f"Processing images in {subfolder}", leave=False):

            parts = image_name.split('_')
            if len(parts) < 4:
                print(f"图片名称格式不正确，跳过: {image_name}")
                continue

            patient_id, eye_code = parts[0], parts[1]
            eye_side = 'left' if eye_code == '21011' else 'right' if eye_code == '21013' else None
            assert eye_side is not None, f"未知的眼睛代码 {eye_code} 在图片 {image_name} 中"

            patient_info = df[df['eid_ckd'] == patient_id]
            src_path = os.path.join(subfolder_path, image_name)

            if patient_info.empty:
                dst_path = os.path.join(no_clinical_dir, image_name)
                os.makedirs(no_clinical_dir, exist_ok=True)
                try:
                    shutil.move(src_path, dst_path)
                except Exception:
                    dst_path = src_path
                missing_images.append(os.path.join(img_dir, os.path.basename(subfolder_path) + '_no_clinical_information', image_name))
                continue

            clinical_info = patient_info.to_dict(orient='records')[0]
            clinical_info = make_serializable(clinical_info)

            if patient_id not in ukb_dataset:
                ukb_dataset[patient_id] = {'left': [], 'right': [],'clinical_info': clinical_info}

            ukb_dataset[patient_id][eye_side].append({
                "image_path": os.path.join(img_dir, os.path.basename(subfolder_path), image_name),
            })

    print("处理图片临床信息总数：", len(ukb_dataset))
    print("前5个图片：")
    for i, (key, value) in enumerate(ukb_dataset.items()):
        if i >= 5:
            break
        print(f"患者ID: {key}, 数据: {value}")

    if save:
        if len(missing_images) > 0:
            missing_images_file = os.path.join(img_dir, 'missing_images.txt')
            with open(missing_images_file, 'w') as txt_file:
                txt_file.write('\n'.join(missing_images))
            print(f"没有找到对应信息的图片路径已保存到 {missing_images_file}")

        output_file = os.path.join(img_dir, 'ukb_dataset.json')
        with open(output_file, 'w') as json_file:
            json.dump(ukb_dataset, json_file, indent=4, ensure_ascii=False)
        print(f"ukb_dataset 已保存到 {output_file}")

    return ukb_dataset, missing_images


# build_ukb_dataset(label_excel_file, img_root_dir, save=True)


# df = pd.read_excel(label_excel_file)

def _map_yesno(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, (bytes, bytearray)):
        try:
            x = x.decode('utf-8')
        except Exception:
            x = str(x)
    if isinstance(x, str):
        s = x.strip()
        if s == '是':
            return 1
        if s == '否':
            return 0
        return s
    if isinstance(x, (np.integer, np.floating)):
        return x.item()
    return x

def _sanitize_for_filename(s):
    if pd.isna(s):
        s = 'NA'
    else:
        s = str(s)
    # 仅保留常见安全字符
    s = s.replace(' ', '_')
    return re.sub(r'[^\w\-\._]', '_', s)

# python
def export_combinations_to_excels(df: pd.DataFrame,
                                  cols=('new_totdepress', 'baseline_depression', 'incident_depression'),
                                  out_dir='.',
                                  save=True):
    """
    按 cols 的所有组合分组，打印每组的样本数（终端以四列对齐），
    将每组对应的行保存为单独的 Excel 文件到 out_dir，并把所有组合统计保存为
    `combination_counts.xlsx`（前三列为组合值，第4列为样本数）。
    返回一个 dict: {组合_tuple: count} 和 保存的统计 DataFrame。
    """
    os.makedirs(out_dir, exist_ok=True)
    ndf = df.copy()

    for c in cols:
        if c not in ndf.columns:
            raise KeyError(f"Column `{c}` not found in DataFrame")
        ndf[c + '_norm'] = ndf[c].apply(_map_yesno)

    # 尝试把 new_totdepress_norm 转为数值
    try:
        ndf['new_totdepress_norm'] = pd.to_numeric(ndf['new_totdepress_norm'], errors='coerce')
    except Exception:
        pass

    norm_cols = [c + '_norm' for c in cols]
    groups = ndf.groupby(norm_cols, dropna=False)

    counts = {}
    stats_rows = []

    def _val_for_output(v):
        # 将 numpy 类型 / pandas NA 转为可存储与显示的 Python 原生类型
        if pd.isna(v):
            return 'NA'
        if isinstance(v, (np.integer,)):
            return int(v)
        if isinstance(v, (np.floating,)):
            # 保留原浮点或转为 Python float
            return float(v)
        # 保证 bytes -> str
        if isinstance(v, (bytes, bytearray)):
            try:
                return v.decode('utf-8')
            except Exception:
                return str(v)
        return v

    for key, grp in groups:
        if not isinstance(key, tuple):
            key = (key,)
        cnt = len(grp)
        counts[key] = cnt

        # 构建用于保存统计表与终端显示的行
        out_vals = [_val_for_output(k) for k in key]
        row = {cols[i]: out_vals[i] for i in range(len(cols))}
        row['count'] = cnt
        stats_rows.append(row)

        # 保存每个组合的子集到单独 Excel（使用原始 df 的行）
        if save and cnt > 0:
            parts = []
            for col_name, val in zip(cols, key):
                parts.append(f"{col_name}_{_sanitize_for_filename(val)}")
            fname = "comb_" + "_".join(parts) + ".xlsx"
            path = os.path.join(out_dir, fname)
            orig_idx = grp.index
            df.loc[orig_idx].to_excel(path, index=False)
            print(f"已保存 {cnt} 行到 `{path}`")

    # 保存统计表
    stats_df = pd.DataFrame(stats_rows, columns=list(cols) + ['count'])
    stats_path = os.path.join(out_dir, 'combination_counts.xlsx')
    stats_df.to_excel(stats_path, index=False)
    print(f"组合统计已保存到 `{stats_path}`")

    # 终端四列对齐打印：计算每列宽度
    headers = list(cols) + ['count']
    rows_as_str = [[str(r[c]) for c in headers] for r in stats_rows]
    col_widths = []
    for i, h in enumerate(headers):
        maxw = max([len(h)] + [len(row[i]) for row in rows_as_str])
        col_widths.append(maxw)
    # 打印表头
    header_line = "  ".join(h.ljust(col_widths[i]) for i, h in enumerate(headers))
    print(header_line)
    print("-" * (sum(col_widths) + 2 * (len(headers) - 1)))
    # 打印每行
    for row in rows_as_str:
        line = "  ".join(row[i].ljust(col_widths[i]) for i in range(len(headers)))
        print(line)

    return counts, stats_df

# # 使用示例：
# df = pd.read_excel(label_excel_file)
# export_combinations_to_excels(df, out_dir=os.path.join(data_split_result_dir,'output_subsets'))

def _get_label_safe(clinical, key):
    """从 clinical_info 中安全读取整数标签（0/1），无法读取返回 None。"""
    if not clinical:
        return None
    v = clinical.get(key)
    if v is None:
        return None
    try:
        return int(v)
    except Exception:
        # 处理字符串形式的 '是'/'否'
        s = str(v).strip()
        if s == '是':
            return 1
        if s == '否':
            return 0
        try:
            return int(float(s))
        except Exception:
            return None

def compute_meta(patient_ids, data_all, total_n=TOTAL_N, denom_remain=None):
    """计算给定 patient_ids 的 meta 信息。"""
    sample_count = len(patient_ids)
    pos = 0
    neg = 0
    left_images = 0
    right_images = 0
    for pid in patient_ids:
        entry = data_all.get(pid, {})
        clinical = entry.get('clinical_info', {})
        lab = _get_label_safe(clinical, 'baseline_depression')
        if lab == 1:
            pos += 1
        elif lab == 0:
            neg += 1
        # image counts
        left_images += len(entry.get('left', []))
        right_images += len(entry.get('right', []))
    denom_remain = denom_remain if denom_remain is not None and denom_remain > 0 else None
    return {
        'sample_count': sample_count,
        'positive_count': pos,
        'negative_count': neg,
        'ratio_of_total': sample_count / total_n if total_n else None,
        'ratio_of_data_remain': sample_count / denom_remain if denom_remain else None,
        'left_image_count': left_images,
        'right_image_count': right_images,
        # 'patient_ids': sorted(patient_ids)
    }

def split_new_totdepress():
    random.seed(SEED)
    output_file = os.path.join(img_root_dir, 'ukb_dataset.json')

    # 1. 载入数据
    with open(output_file, 'r', encoding='utf-8') as f:
        data_all = json.load(f)

    all_ids = sorted(data_all.keys())
    denom_remain = len(all_ids)

    # 2. 计算 val/test 大小（各占 TOTAL_N 的 10%），若超出则按 denom_remain 调整
    val_n = int(round(TOTAL_N * 0.10))
    test_n = int(round(TOTAL_N * 0.10))
    if val_n + test_n > denom_remain:
        # 简单回退：以 denom_remain 的 10% 为基准（向下取整）
        val_n = denom_remain // 10
        test_n = denom_remain // 10

    # 3. 随机抽取 val 和 test（互斥），其余为 train
    pool = all_ids.copy()
    random.shuffle(pool)
    val_ids = set(pool[:val_n])
    test_ids = set(pool[val_n:val_n + test_n])
    train_ids = [pid for pid in all_ids if pid not in val_ids and pid not in test_ids]

    # 4. 构建拆分条目（将 new_totdepress 转为 gt_label）
    def build_split_dict(id_list):
        meta = compute_meta(sorted(id_list), data_all, total_n=TOTAL_N, denom_remain=denom_remain)
        patients = {}
        for pid in sorted(id_list):
            entry = data_all.get(pid, {})
            clinical = entry.get('clinical_info', {})
            lab = _get_label_safe(clinical, 'new_totdepress')
            gt = 1 if lab == 1 else 0
            pat = {'clinical_info': clinical, 'left': [], 'right': []}
            for side in ('left', 'right'):
                imgs = []
                for img in entry.get(side, []):
                    new_img = dict(img)
                    new_img['gt_label'] = gt
                    imgs.append(new_img)
                pat[side] = imgs
            patients[pid] = pat
        return {'meta': meta, 'data': patients}

    split_obj = {
        'train': build_split_dict(train_ids),
        'val': build_split_dict(sorted(val_ids)),
        'test': build_split_dict(sorted(test_ids)),
    }

    # 5. 保存拆分 JSON
    base_dir = os.path.dirname(data_split_result_dir)
    os.makedirs(base_dir, exist_ok=True)
    out = os.path.join(base_dir, 'new_totdepress_data_split.json')
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(split_obj, f, ensure_ascii=False, indent=2)
    print(f"已保存拆分到 ` {out} `")

    # 6. 生成统计表并写入 Excel
    def make_stats_df(split_obj, denom_remain, total_n):
        rows = []
        keys_order = ['train', 'val', 'test']
        for k in keys_order:
            if k not in split_obj:
                continue
            meta = split_obj[k]['meta']
            rows.append({
                'split': k,
                'sample_count': meta.get('sample_count', 0),
                'positive_count': meta.get('positive_count', 0),
                'negative_count': meta.get('negative_count', 0),
                'left_image_count': meta.get('left_image_count', 0),
                'right_image_count': meta.get('right_image_count', 0),
                'ratio_of_total': meta.get('ratio_of_total', 0.0) if meta.get('ratio_of_total') is not None else 0.0,
                'ratio_of_data_remain': meta.get('ratio_of_data_remain', 0.0) if meta.get('ratio_of_data_remain') is not None else 0.0,
            })
        df = pd.DataFrame(rows, columns=[
            'split', 'sample_count', 'positive_count', 'negative_count',
            'left_image_count', 'right_image_count', 'ratio_of_total', 'ratio_of_data_remain'
        ])
        total_sample = int(df['sample_count'].sum()) if not df.empty else 0
        total_pos = int(df['positive_count'].sum()) if not df.empty else 0
        total_neg = int(df['negative_count'].sum()) if not df.empty else 0
        total_left = int(df['left_image_count'].sum()) if not df.empty else 0
        total_right = int(df['right_image_count'].sum()) if not df.empty else 0
        ratio_of_total = total_sample / total_n if total_n else 0.0
        ratio_of_data_remain = total_sample / denom_remain if denom_remain else 0.0
        total_row = {
            'split': 'total',
            'sample_count': total_sample,
            'positive_count': total_pos,
            'negative_count': total_neg,
            'left_image_count': total_left,
            'right_image_count': total_right,
            'ratio_of_total': ratio_of_total,
            'ratio_of_data_remain': ratio_of_data_remain
        }
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        return df

    stats = make_stats_df(split_obj, denom_remain, TOTAL_N)
    stats_xlsx = os.path.join(base_dir, 'new_totdepress_splits_stats.xlsx')
    with pd.ExcelWriter(stats_xlsx, engine='openpyxl') as writer:
        stats.to_excel(writer, sheet_name='splits', index=False)

    # 格式化百分比列
    wb = load_workbook(stats_xlsx)
    if 'splits' in wb.sheetnames:
        ws = wb['splits']
        header = [cell.value for cell in ws[1]]
        for col_name in ['ratio_of_total', 'ratio_of_data_remain']:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is None:
                        continue
                    try:
                        cell.value = float(cell.value)
                    except Exception:
                        try:
                            cell.value = float(str(cell.value))
                        except Exception:
                            continue
                    cell.number_format = '0.0%'
    wb.save(stats_xlsx)
    print(f"统计表已保存到 ` {stats_xlsx} `")


split_new_totdepress()

def split_baseline_depression():
    random.seed(SEED)
    output_file = os.path.join(img_root_dir, 'ukb_dataset.json')

    # 1. 载入数据
    with open(output_file, 'r', encoding='utf-8') as f:
        data_all = json.load(f)

    all_ids = sorted(data_all.keys())
    # 2. 找出有争议样本 baseline_depression==1 且 new_totdepress==0
    ambiguous_ids = []
    for pid, entry in data_all.items():
        clin = entry.get('clinical_info', {})
        b = _get_label_safe(clin, 'baseline_depression')
        n = _get_label_safe(clin, 'new_totdepress')
        if b == 1 and n == 0:
            ambiguous_ids.append(pid)

    # base_dir = os.path.dirname(output_file)
    base_dir = os.path.dirname(data_split_result_dir)
    os.makedirs(base_dir, exist_ok=True)
    # 保存 ambiguous 全条目（保留单独文件）

    ambiguous_out = os.path.join(base_dir, 'baseline_depression_ambiguous.json')
    with open(ambiguous_out, 'w', encoding='utf-8') as f:
        amb_dict = {pid: data_all[pid] for pid in ambiguous_ids}
        json.dump(amb_dict, f, ensure_ascii=False, indent=2)
    print(f"已保存 {len(ambiguous_ids)} 个有争议患者到 `{ambiguous_out}`")

    # 3. 构建 data_remain（排除 ambiguous）
    ambiguous_set = set(ambiguous_ids)
    data_remain_ids = [pid for pid in all_ids if pid not in ambiguous_set]
    denom_remain = len(data_remain_ids)

    # 4. 计算 val/test 样本数（各占 TOTAL_N 的 10%，四舍五入）
    val_n = int(round(TOTAL_N * 0.10))
    test_n = int(round(TOTAL_N * 0.10))
    if val_n + test_n > denom_remain:
        raise ValueError("要求的 val/test 大小超过了有效样本数 data_remain，请调整比例或处理 ambiguous。")

    # 5. 从 data_remain 随机抽取 val 和 test（互斥）
    pool = data_remain_ids.copy()
    random.shuffle(pool)
    val_ids = set(pool[:val_n])
    test_ids = set(pool[val_n:val_n + test_n])

    # 6. 构建两种 train：
    train_remain_ids = [pid for pid in data_remain_ids if pid not in val_ids and pid not in test_ids]  # 第一种：仅来自 data_remain
    train_with_ambiguous_ids = [pid for pid in all_ids if pid not in val_ids and pid not in test_ids]  # 第二种：来自 data_all（含 ambiguous）

    # 7. 构建包含完整患者信息的拆分结构并保存 JSON

    # def build_split_dict(id_list):
    #     meta = compute_meta(sorted(id_list), data_all, total_n=TOTAL_N, denom_remain=denom_remain)
    #     patients = {pid: data_all[pid] for pid in sorted(id_list)}
    #     return {'meta': meta, 'data': patients}

    def build_split_dict(id_list):
        meta = compute_meta(sorted(id_list), data_all, total_n=TOTAL_N, denom_remain=denom_remain)
        patients = {}
        for pid in sorted(id_list):
            entry = data_all.get(pid, {})
            clinical = entry.get('clinical_info', {})
            lab = _get_label_safe(clinical, 'baseline_depression')
            gt = 1 if lab == 1 else 0
            # 构造不修改原始 data_all 的新条目
            pat = {'clinical_info': clinical, 'left': [], 'right': []}
            for side in ('left', 'right'):
                imgs = []
                for img in entry.get(side, []):
                    new_img = dict(img)  # 复制图片字典
                    new_img['gt_label'] = gt
                    imgs.append(new_img)
                pat[side] = imgs
            patients[pid] = pat
        return {'meta': meta, 'data': patients}

    # 在最终 json 中加入 ambiguous 键（含 meta 和 data）
    split1_obj = {
        'train': build_split_dict(train_remain_ids),
        'val': build_split_dict(sorted(val_ids)),
        'test': build_split_dict(sorted(test_ids)),
        'ambiguous': build_split_dict(sorted(ambiguous_ids)),
    }
    out1 = os.path.join(base_dir, 'baseline_depression_data_split_no_ambiguous.json')
    with open(out1, 'w', encoding='utf-8') as f:
        json.dump(split1_obj, f, ensure_ascii=False, indent=2)
    print(f"已保存拆分（不含 ambiguous in train，但含 `ambiguous` 键）到 `{out1}`")

    # 第二种：含 ambiguous（ambiguous 已包含在 train），并且也保留 ambiguous 键以供统计
    split2_obj = {
        'train': build_split_dict(train_with_ambiguous_ids),
        'val': build_split_dict(sorted(val_ids)),
        'test': build_split_dict(sorted(test_ids)),
    }
    out2 = os.path.join(base_dir, 'baseline_depression_data_split.json')
    with open(out2, 'w', encoding='utf-8') as f:
        json.dump(split2_obj, f, ensure_ascii=False, indent=2)
    print(f"已保存拆分（含 ambiguous in train，且含 `ambiguous` 键）到 `{out2}`")

    # 8. 生成用于写入 Excel 的统计表（包含 ambiguous，并增加 total 汇总行）
    def make_stats_df(split_obj, denom_remain, total_n):
        rows = []
        keys_order = ['train', 'val', 'test', 'ambiguous']
        for k in keys_order:
            if k not in split_obj:
                continue
            meta = split_obj[k]['meta']
            rows.append({
                'split': k,
                'sample_count': meta.get('sample_count', 0),
                'positive_count': meta.get('positive_count', 0),
                'negative_count': meta.get('negative_count', 0),
                'left_image_count': meta.get('left_image_count', 0),
                'right_image_count': meta.get('right_image_count', 0),
                # 保存为小数（例如 0.10），后续写 Excel 时应用百分比格式
                'ratio_of_total': meta.get('ratio_of_total', 0.0) if meta.get('ratio_of_total') is not None else 0.0,
                'ratio_of_data_remain': meta.get('ratio_of_data_remain', 0.0) if meta.get('ratio_of_data_remain') is not None else 0.0,
            })
        df = pd.DataFrame(rows, columns=[
            'split', 'sample_count', 'positive_count', 'negative_count',
            'left_image_count', 'right_image_count', 'ratio_of_total', 'ratio_of_data_remain'
        ])
        # 汇总行：对计数列求和，比例按合计 sample_count 重新计算（基于 total_n 和 denom_remain）
        total_sample = int(df['sample_count'].sum()) if not df.empty else 0
        total_pos = int(df['positive_count'].sum()) if not df.empty else 0
        total_neg = int(df['negative_count'].sum()) if not df.empty else 0
        total_left = int(df['left_image_count'].sum()) if not df.empty else 0
        total_right = int(df['right_image_count'].sum()) if not df.empty else 0
        ratio_of_total = total_sample / total_n if total_n else 0.0
        ratio_of_data_remain = total_sample / denom_remain if denom_remain else 0.0
        total_row = {
            'split': 'total',
            'sample_count': total_sample,
            'positive_count': total_pos,
            'negative_count': total_neg,
            'left_image_count': total_left,
            'right_image_count': total_right,
            'ratio_of_total': ratio_of_total,
            'ratio_of_data_remain': ratio_of_data_remain
        }
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        return df

    stats1 = make_stats_df(split1_obj, denom_remain, TOTAL_N)
    stats2 = make_stats_df(split2_obj, denom_remain, TOTAL_N)

    # 9. 保存到同一个 Excel 文件，两个 sheet，并把比例列格式化为百分数（1 位小数）
    stats_xlsx = os.path.join(base_dir, 'baseline_depression_splits_stats.xlsx')
    with pd.ExcelWriter(stats_xlsx, engine='openpyxl') as writer:
        stats1.to_excel(writer, sheet_name='without_ambiguous', index=False)
        stats2.to_excel(writer, sheet_name='with_ambiguous', index=False)

    # 使用 openpyxl 设置百分比格式（保留一位小数）
    wb = load_workbook(stats_xlsx)
    for sheet_name in ['without_ambiguous', 'with_ambiguous']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        for col_name in ['ratio_of_total', 'ratio_of_data_remain']:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is None:
                        continue
                    # 确保单元格为 float 值，然后设置为百分比格式（1 位小数）
                    try:
                        cell.value = float(cell.value)
                    except Exception:
                        # 如果单元格是字符串 "0.1" 等也尝试转换，失败则跳过
                        try:
                            cell.value = float(str(cell.value))
                        except Exception:
                            continue
                    cell.number_format = '0.0%'

    wb.save(stats_xlsx)
    print(f"统计表已保存到 `{stats_xlsx}`（两个 sheet: without_ambiguous, with_ambiguous；含 ambiguous 行及 total 汇总）")


# 调用
# split_baseline_depression()

# python
def split_incident_depression():
    random.seed(SEED)
    output_file = os.path.join(img_root_dir, 'ukb_dataset.json')
    with open(output_file, 'r', encoding='utf-8') as f:
        data_all = json.load(f)

    all_ids = sorted(data_all.keys())

    # 1. 彻底排除 baseline_depression == 1
    baseline1_ids = [pid for pid in all_ids
                     if _get_label_safe(data_all[pid].get('clinical_info', {}), 'baseline_depression') == 1]
    baseline1_set = set(baseline1_ids)
    all_ids_filtered = [pid for pid in all_ids if pid not in baseline1_set]
    baseline1_count = len(baseline1_ids)
    print(f"排除 baseline_depression==1 的样本数: {baseline1_count}")
    # 总样本数 N = TOTAL_N - baseline1_count (依据要求)
    N = TOTAL_N - baseline1_count

    # 2. 找出有争议样本： incident_depression==1 且 new_totdepress==0 （在已排除 baseline1 的集合内）
    ambiguous_ids = []
    for pid in all_ids_filtered:
        clin = data_all[pid].get('clinical_info', {})
        inc = _get_label_safe(clin, 'incident_depression')
        n = _get_label_safe(clin, 'new_totdepress')
        if inc == 1 and n == 0:
            ambiguous_ids.append(pid)

    # 保存 ambiguous 全条目（文件名中的 `_ambiguous` 改为 `_all`）
    base_dir = os.path.dirname(data_split_result_dir)
    os.makedirs(base_dir, exist_ok=True)

    ambiguous_out = os.path.join(base_dir, 'incident_depression_ambiguous.json')
    with open(ambiguous_out, 'w', encoding='utf-8') as f:
        amb_dict = {pid: data_all[pid] for pid in ambiguous_ids}
        json.dump(amb_dict, f, ensure_ascii=False, indent=2)
    print(f"已保存 {len(ambiguous_ids)} 个有争议患者到 `{ambiguous_out}`")

    ambiguous_set = set(ambiguous_ids)

    # 3. data_remain：排除 ambiguous（并且已经排除了 baseline1）
    data_remain_ids = [pid for pid in all_ids_filtered if pid not in ambiguous_set]
    denom_remain = len(data_remain_ids)

    # 4. 计算 val/test 大小（各占 N 的 10%）
    val_n = int(round(N * 0.10))
    test_n = int(round(N * 0.10))
    if val_n + test_n > denom_remain:
        raise ValueError("要求的 val/test 大小超过了有效样本数 data_remain，请调整比例或处理 ambiguous。")

    # 5. 随机抽取 val 和 test（互斥）
    pool = data_remain_ids.copy()
    random.shuffle(pool)
    val_ids = set(pool[:val_n])
    test_ids = set(pool[val_n:val_n + test_n])

    # 6. 构建两种 train：
    train_remain_ids = [pid for pid in data_remain_ids if pid not in val_ids and pid not in test_ids]
    # 第二种：基于 all_ids_filtered（含 ambiguous），但仍排除了 baseline_depression==1
    train_with_ambiguous_ids = [pid for pid in all_ids_filtered if pid not in val_ids and pid not in test_ids]

    # 内部：按指定标签计算 meta（用于 incident_depression）
    def compute_meta_label(patient_ids, data_all, label_key='incident_depression', total_n=N, denom_remain=denom_remain):
        sample_count = len(patient_ids)
        pos = 0
        neg = 0
        left_images = 0
        right_images = 0
        for pid in patient_ids:
            entry = data_all.get(pid, {})
            clinical = entry.get('clinical_info', {})
            lab = _get_label_safe(clinical, label_key)
            if lab == 1:
                pos += 1
            elif lab == 0:
                neg += 1
            left_images += len(entry.get('left', []))
            right_images += len(entry.get('right', []))
        denom = denom_remain if denom_remain and denom_remain > 0 else None
        return {
            'sample_count': sample_count,
            'positive_count': pos,
            'negative_count': neg,
            'ratio_of_total': sample_count / total_n if total_n else None,
            'ratio_of_data_remain': sample_count / denom if denom else None,
            'left_image_count': left_images,
            'right_image_count': right_images
        }

    # 7. 构建包含完整患者信息的拆分结构并保存 JSON
    # def build_split_item(id_list):
    #     meta = compute_meta_label(sorted(id_list), data_all, label_key='incident_depression', total_n=N, denom_remain=denom_remain)
    #     patients = {pid: data_all[pid] for pid in sorted(id_list)}
    #     return {'meta': meta, 'data': patients}
    def build_split_item(id_list):
        meta = compute_meta_label(sorted(id_list), data_all, label_key='incident_depression', total_n=N,
                                  denom_remain=denom_remain)
        patients = {}
        for pid in sorted(id_list):
            entry = data_all.get(pid, {})
            clinical = entry.get('clinical_info', {})
            lab = _get_label_safe(clinical, 'incident_depression')
            gt = 1 if lab == 1 else 0
            # 构造不修改原始 data_all 的新条目
            pat = {'clinical_info': clinical, 'left': [], 'right': []}
            for side in ('left', 'right'):
                imgs = []
                for img in entry.get(side, []):
                    new_img = dict(img)  # 复制图片字典
                    new_img['gt_label'] = gt
                    imgs.append(new_img)
                pat[side] = imgs
            patients[pid] = pat
        return {'meta': meta, 'data': patients}

    # 第一种：不含 ambiguous（但保留 ambiguous 在文件的单独键），并增加 excluded 键（baseline==1 的完整数据，带 meta）
    split1_obj = {
        'train': build_split_item(train_remain_ids),
        'val': build_split_item(sorted(val_ids)),
        'test': build_split_item(sorted(test_ids)),
        'excluded': build_split_item(sorted(baseline1_ids)),
        'ambiguous': build_split_item(sorted(ambiguous_ids)),
    }
    out1 = os.path.join(base_dir, 'incident_depression_data_split_no_ambiguous.json')
    with open(out1, 'w', encoding='utf-8') as f:
        json.dump(split1_obj, f, ensure_ascii=False, indent=2)
    print(f"已保存拆分（不含 ambiguous，含 `excluded` 和 `ambiguous`）到 `{out1}`")

    # 第二种：含 ambiguous（ambiguous 已包含在 train），但仍保留 excluded 键
    split2_obj = {
        'train': build_split_item(train_with_ambiguous_ids),
        'val': build_split_item(sorted(val_ids)),
        'test': build_split_item(sorted(test_ids)),
        'excluded': build_split_item(sorted(baseline1_ids)),
    }
    out2 = os.path.join(base_dir, 'incident_depression_data_split.json')  # 将文件名中的 _ambiguous 替换为 _all
    with open(out2, 'w', encoding='utf-8') as f:
        json.dump(split2_obj, f, ensure_ascii=False, indent=2)
    print(f"已保存拆分（含 ambiguous in train，含 `excluded`）到 `{out2}`")

    # 8. 生成统计表（包含 total 汇总行），把所有可用的键（train/val/test/ambiguous/excluded）都统计并汇总
    def make_stats_df(split_obj, denom_remain, total_n):
        rows = []
        # 固定输出顺序
        keys_order = ['train', 'val', 'test', 'ambiguous', 'excluded']
        for k in keys_order:
            if k not in split_obj:
                continue
            meta = split_obj[k]['meta']
            rows.append({
                'split': k,
                'sample_count': meta['sample_count'],
                'positive_count': meta['positive_count'],
                'negative_count': meta['negative_count'],
                'left_image_count': meta['left_image_count'],
                'right_image_count': meta['right_image_count'],
                # 保存为小数（例如 0.10），后续写 Excel 时应用百分比格式
                'ratio_of_total': meta['ratio_of_total'] if meta.get('ratio_of_total') is not None else 0.0,
                'ratio_of_data_remain': meta['ratio_of_data_remain'] if meta.get('ratio_of_data_remain') is not None else 0.0,
            })
        df = pd.DataFrame(rows, columns=[
            'split', 'sample_count', 'positive_count', 'negative_count',
            'left_image_count', 'right_image_count', 'ratio_of_total', 'ratio_of_data_remain'
        ])
        # 汇总行：对计数列求和，比例按合计 sample_count 重新计算（基于 total_n 和 denom_remain）
        total_sample = int(df['sample_count'].sum()) if not df.empty else 0
        total_pos = int(df['positive_count'].sum()) if not df.empty else 0
        total_neg = int(df['negative_count'].sum()) if not df.empty else 0
        total_left = int(df['left_image_count'].sum()) if not df.empty else 0
        total_right = int(df['right_image_count'].sum()) if not df.empty else 0
        ratio_of_total = total_sample / total_n if total_n else 0.0
        ratio_of_data_remain = total_sample / denom_remain if denom_remain else 0.0
        total_row = {
            'split': 'total',
            'sample_count': total_sample,
            'positive_count': total_pos,
            'negative_count': total_neg,
            'left_image_count': total_left,
            'right_image_count': total_right,
            'ratio_of_total': ratio_of_total,
            'ratio_of_data_remain': ratio_of_data_remain
        }
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        return df

    stats1 = make_stats_df(split1_obj, denom_remain, N)
    stats2 = make_stats_df(split2_obj, denom_remain, N)

    # 9. 保存到同一个 Excel 文件，两个 sheet，并把比例列格式化为百分数（1 位小数）
    stats_xlsx = os.path.join(base_dir, 'incident_depression_splits_stats.xlsx')
    with pd.ExcelWriter(stats_xlsx, engine='openpyxl') as writer:
        # sheet 名称中的 ambiguous -> all
        stats1.to_excel(writer, sheet_name='without_ambiguous', index=False)
        stats2.to_excel(writer, sheet_name='with_ambiguous', index=False)

    wb = load_workbook(stats_xlsx)
    for sheet_name in ['without_ambiguous', 'with_ambiguous']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        for col_name in ['ratio_of_total', 'ratio_of_data_remain']:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is None:
                        continue
                    # 尝试确保为浮点数（pandas 写入通常已为浮点）
                    try:
                        cell.value = float(cell.value)
                    except Exception:
                        # 如果不是数字，跳过转换
                        pass
                    cell.number_format = '0.0%'

    wb.save(stats_xlsx)
    print(f"统计表已保存到 `{stats_xlsx}`（两个 sheet: without_all, with_all）")
# 调用示例：
# split_incident_depression()



